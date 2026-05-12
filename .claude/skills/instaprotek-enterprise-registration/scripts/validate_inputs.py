"""Validate, clean, and cross-check the bulk registration file against the PO.

Responsibilities:
- Load CSV or XLSX into an in-memory row list (headers + rows).
- Auto-fix: trim whitespace, normalize Yes/No, title-case Manufacturer.
- Cross-validate row count against PO quantity.
- Cross-validate end-user name against PO end-user name (best-effort fuzzy match).
- Identify rows whose Model Number is SKU-shaped — flagged for resolve_model_names.
- Write the corrected file under the original filename, and the untouched original
  alongside with an `.original` suffix.

The module never touches the CRM. All Brand-menu reconciliation is performed by
resolve_model_names.py.
"""
from __future__ import annotations

import csv
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from utils import (
    collapse_ws,
    looks_like_sku,
    normalize_name,
    normalize_yes_no,
    title_case_manufacturer,
)


REQUIRED_HEADERS = [
    "First Name",
    "Last Name",
    "Phone Number",
    "Email",
    "Street Address",
    "City",
    "State",
    "Zip Code",
    "Serial Number",
    "Manufacturer",
    "Model Number",
    "Delivery Date",
    "Purchase Date",
    "Purchase Price",
    "New? (Yes/No)",
]


@dataclass
class RegistrationFile:
    source_path: Path
    headers: list[str]
    rows: list[dict[str, Any]]
    extension: str  # ".csv" or ".xlsx"

    @property
    def row_count(self) -> int:
        return len(self.rows)


@dataclass
class ValidationResult:
    cleaned: RegistrationFile
    fixes_applied: list[dict[str, Any]] = field(default_factory=list)
    rows_needing_model_resolution: list[int] = field(default_factory=list)  # 1-based data row numbers
    cross_validation: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


# ---- Loaders --------------------------------------------------------------


def load_registration_file(path: Path) -> RegistrationFile:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _load_csv(path)
    if ext in (".xlsx", ".xls"):
        return _load_xlsx(path)
    raise ValueError(f"Unsupported registration file extension: {ext}")


def _load_csv(path: Path) -> RegistrationFile:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows_raw = list(reader)
    if not rows_raw:
        raise ValueError(f"CSV is empty: {path}")
    headers = [h.strip() for h in rows_raw[0]]
    rows: list[dict[str, Any]] = []
    for raw in rows_raw[1:]:
        if not any((cell or "").strip() for cell in raw):
            continue
        row = {h: (raw[i] if i < len(raw) else "") for i, h in enumerate(headers)}
        rows.append(row)
    return RegistrationFile(source_path=path, headers=headers, rows=rows, extension=".csv")


def _load_xlsx(path: Path) -> RegistrationFile:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first_row = next(rows_iter)
    except StopIteration:
        raise ValueError(f"XLSX is empty: {path}")

    # Determine how many columns are real headers (trim trailing Nones).
    headers: list[str] = []
    for value in first_row:
        if value is None:
            break
        headers.append(str(value).strip())

    rows: list[dict[str, Any]] = []
    for raw in rows_iter:
        # Skip rows that are entirely blank in the header columns
        sliced = list(raw[: len(headers)])
        if not any(_nonempty(v) for v in sliced):
            continue
        row = {h: sliced[i] if i < len(sliced) else None for i, h in enumerate(headers)}
        rows.append(row)
    return RegistrationFile(source_path=path, headers=headers, rows=rows, extension=path.suffix.lower())


def _nonempty(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str) and not v.strip():
        return False
    return True


# ---- Validation ---------------------------------------------------------


def validate_and_clean(
    reg: RegistrationFile,
    po,
    *,
    sku_pattern: str | None = None,
    auto_trim: bool = True,
    auto_yes_no: bool = True,
    auto_titlecase_manufacturer: bool = True,
) -> ValidationResult:
    result = ValidationResult(cleaned=reg)

    _check_required_headers(reg, result)
    if result.has_errors:
        return result

    _clean_rows(
        reg,
        result,
        auto_trim=auto_trim,
        auto_yes_no=auto_yes_no,
        auto_titlecase_manufacturer=auto_titlecase_manufacturer,
        sku_pattern=sku_pattern,
    )
    _cross_validate(reg, po, result)
    return result


def _check_required_headers(reg: RegistrationFile, result: ValidationResult) -> None:
    missing = [h for h in REQUIRED_HEADERS if h not in reg.headers]
    if missing:
        result.errors.append(
            f"Registration file is missing required columns: {missing}. Found: {reg.headers}"
        )


def _clean_rows(
    reg: RegistrationFile,
    result: ValidationResult,
    *,
    auto_trim: bool,
    auto_yes_no: bool,
    auto_titlecase_manufacturer: bool,
    sku_pattern: str | None,
) -> None:
    for idx, row in enumerate(reg.rows, start=1):
        for header in reg.headers:
            original = row.get(header)
            new_value = original
            if auto_trim and isinstance(new_value, str):
                trimmed = collapse_ws(new_value)
                if trimmed != new_value:
                    result.fixes_applied.append(
                        {
                            "row": idx,
                            "column": header,
                            "fix": "trim_whitespace",
                            "before": new_value,
                            "after": trimmed,
                        }
                    )
                    new_value = trimmed
            if header == "New? (Yes/No)" and auto_yes_no:
                normalized = normalize_yes_no(new_value)
                if normalized != new_value:
                    result.fixes_applied.append(
                        {
                            "row": idx,
                            "column": header,
                            "fix": "normalize_yes_no",
                            "before": new_value,
                            "after": normalized,
                        }
                    )
                    new_value = normalized
            if header == "Manufacturer" and auto_titlecase_manufacturer and isinstance(new_value, str):
                tc = title_case_manufacturer(new_value)
                if tc != new_value:
                    result.fixes_applied.append(
                        {
                            "row": idx,
                            "column": header,
                            "fix": "title_case_manufacturer",
                            "before": new_value,
                            "after": tc,
                        }
                    )
                    new_value = tc
            row[header] = new_value

        # SKU detection on the (possibly-updated) Model Number
        model_number = row.get("Model Number")
        if looks_like_sku(model_number, sku_pattern):
            result.rows_needing_model_resolution.append(idx)


def _cross_validate(reg: RegistrationFile, po, result: ValidationResult) -> None:
    cv: dict[str, Any] = {
        "csv_row_count": reg.row_count,
        "po_quantity": getattr(po, "quantity", None),
        "row_count_match": None,
        "csv_end_user_name": None,
        "po_end_user_name": getattr(getattr(po, "end_user", None), "name", None),
        "end_user_match": None,
    }

    # Row count vs PO quantity
    if po.quantity is None:
        result.warnings.append(
            "PO quantity could not be parsed — skipping row count cross-validation."
        )
    else:
        cv["row_count_match"] = reg.row_count == po.quantity
        if not cv["row_count_match"]:
            result.errors.append(
                f"Registration row count ({reg.row_count}) does not match PO quantity ({po.quantity})."
            )

    # End-user name cross-validation: first row's First+Last vs PO end-user name
    if reg.rows:
        first = reg.rows[0]
        csv_name = collapse_ws(f"{first.get('First Name', '')} {first.get('Last Name', '')}").strip()
        cv["csv_end_user_name"] = csv_name
        po_name = cv["po_end_user_name"]
        if csv_name and po_name:
            cv["end_user_match"] = normalize_name(csv_name) == normalize_name(po_name)
            if not cv["end_user_match"]:
                result.errors.append(
                    f"End-user mismatch: registration first row is '{csv_name}', PO is '{po_name}'."
                )
        else:
            result.warnings.append(
                f"End-user cross-validation skipped (csv_name='{csv_name}', po_name='{po_name}')."
            )

    result.cross_validation = cv


# ---- Writers --------------------------------------------------------------


def write_cleaned_and_preserved(
    reg: RegistrationFile,
    destination_folder: Path,
    original_source: Path,
) -> tuple[Path, Path]:
    """Write the cleaned registration under the original filename to `destination_folder`,
    and copy the untouched original alongside with an `.original` suffix inserted before
    the file extension. Returns (cleaned_path, original_copy_path).
    """
    destination_folder.mkdir(parents=True, exist_ok=True)
    original_name = original_source.name
    stem = original_source.stem
    ext = original_source.suffix
    cleaned_path = destination_folder / original_name
    original_copy_path = destination_folder / f"{stem}.original{ext}"

    # Preserve the original input byte-for-byte
    shutil.copy2(original_source, original_copy_path)

    # Write the cleaned version using the same extension
    if ext.lower() == ".csv":
        _write_csv(reg, cleaned_path)
    elif ext.lower() in (".xlsx", ".xls"):
        _write_xlsx(reg, cleaned_path)
    else:
        raise ValueError(f"Unsupported registration file extension: {ext}")

    return cleaned_path, original_copy_path


def _write_csv(reg: RegistrationFile, path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(reg.headers)
        for row in reg.rows:
            writer.writerow([_cell_for_csv(row.get(h)) for h in reg.headers])


def _cell_for_csv(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _write_xlsx(reg: RegistrationFile, path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(reg.headers)
    for row in reg.rows:
        ws.append([row.get(h) for h in reg.headers])
    wb.save(str(path))


if __name__ == "__main__":  # pragma: no cover — manual smoke test
    import json
    import sys

    if len(sys.argv) != 3:
        print("usage: validate_inputs.py <registration.csv|xlsx> <po.pdf>")
        sys.exit(2)
    from parse_po import parse_purchase_order

    reg = load_registration_file(Path(sys.argv[1]))
    po = parse_purchase_order(Path(sys.argv[2]))
    result = validate_and_clean(reg, po)
    out = {
        "row_count": reg.row_count,
        "headers": reg.headers,
        "fixes": result.fixes_applied,
        "needs_resolution_rows": result.rows_needing_model_resolution,
        "cross_validation": result.cross_validation,
        "errors": result.errors,
        "warnings": result.warnings,
    }
    print(json.dumps(out, indent=2, default=str))
