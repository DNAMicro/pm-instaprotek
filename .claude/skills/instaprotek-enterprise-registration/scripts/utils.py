"""Shared helpers for the InstaProtek enterprise registration skill."""
from __future__ import annotations

import json
import logging
import re
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


# ---- Paths ----------------------------------------------------------------

SKILL_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = SKILL_DIR.parent.parent.parent  # .claude/skills/<name>/ -> project root
CONFIG_DIR = SKILL_DIR / "config"
AUTH_DIR = SKILL_DIR / ".auth"
INPUT_DIR = PROJECT_ROOT / "input"
PROCESSED_DIR = PROJECT_ROOT / "processed"
FAILURES_DIR = PROCESSED_DIR / "failures"
CREDENTIALS_PATH = PROJECT_ROOT / "credentials.json"
REFERENCES_DIR = PROJECT_ROOT / "references"
_SKU_LIST_FILENAME = "instaProtek Product SKU List.xlsx"


# ---- Logging --------------------------------------------------------------


def setup_logger(log_file: Path | None = None, verbose: bool = False) -> logging.Logger:
    """Create a structured logger that writes to a run log file and (optionally) stdout."""
    logger = logging.getLogger("instaprotek")
    logger.setLevel(logging.DEBUG)
    # Wipe handlers in case the orchestrator is re-entrant
    logger.handlers.clear()

    fmt = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%dT%H:%M:%S%z",
    )

    if log_file is not None:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        fh = logging.FileHandler(log_file, encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(fmt)
        logger.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.DEBUG if verbose else logging.INFO)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    return logger


# ---- JSON IO --------------------------------------------------------------


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, sort_keys=False, default=_json_default)


def _json_default(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    raise TypeError(f"Object of type {type(value)} is not JSON serializable")


# ---- Dependency check ----------------------------------------------------


REQUIRED_PACKAGES = {
    "playwright": "playwright",
    "pdfplumber": "pdfplumber",
    "openpyxl": "openpyxl",
    "requests": "requests",
}


def ensure_dependencies(logger: logging.Logger) -> None:
    """Install any missing Python packages and the Chromium browser."""
    missing = []
    for module_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            __import__(module_name)
        except ImportError:
            missing.append(pip_name)

    if missing:
        logger.info("Installing missing Python packages: %s", ", ".join(missing))
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--break-system-packages", *missing]
        )

    # Ensure Chromium is installed for Playwright. This is idempotent.
    try:
        import playwright  # noqa: F401
        logger.info("Ensuring Playwright Chromium browser is installed")
        subprocess.check_call(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.STDOUT,
        )
    except Exception as exc:  # pragma: no cover — best-effort
        logger.warning("Playwright install command failed: %s", exc)


# ---- File discovery -------------------------------------------------------


PO_PDF_GLOB = "*.pdf"
REG_FILE_GLOBS = ("*.csv", "*.xlsx", "*.xls")
SOP_KEYWORDS = ("Instaprotek_Enterprise_Sale_Reporting", "Enterprise_Sale_Reporting")


@dataclass
class DiscoveredFiles:
    po_pdf: Path
    registration_file: Path
    ignored: list[Path] = field(default_factory=list)


def discover_input_files(input_dir: Path = INPUT_DIR) -> DiscoveredFiles:
    """Return the PO PDF and the registration CSV/XLSX from input/.

    Raises ValueError if 0 or >1 of either is present (ignoring the SOP file itself).
    """
    if not input_dir.exists():
        raise FileNotFoundError(f"Input folder does not exist: {input_dir}")

    pdfs: list[Path] = []
    reg_files: list[Path] = []
    ignored: list[Path] = []

    for entry in sorted(input_dir.iterdir()):
        if entry.name.startswith(".") or entry.is_dir():
            continue
        # Office lock files (~$...) and other tilde-prefixed temp files
        if entry.name.startswith("~$") or entry.name.startswith("~"):
            ignored.append(entry)
            continue
        if any(keyword.lower() in entry.name.lower() for keyword in SOP_KEYWORDS):
            ignored.append(entry)
            continue
        suffix = entry.suffix.lower()
        if suffix == ".pdf":
            pdfs.append(entry)
        elif suffix in (".csv", ".xlsx", ".xls"):
            reg_files.append(entry)
        else:
            ignored.append(entry)

    if len(pdfs) == 0:
        raise ValueError("No Purchase Order PDF found in input/.")
    if len(pdfs) > 1:
        raise ValueError(
            f"Multiple PDFs found in input/: {[p.name for p in pdfs]}. Leave exactly one PO PDF."
        )
    if len(reg_files) == 0:
        raise ValueError("No CSV or XLSX registration file found in input/.")
    if len(reg_files) > 1:
        raise ValueError(
            f"Multiple registration files found in input/: {[p.name for p in reg_files]}. Leave exactly one."
        )

    return DiscoveredFiles(po_pdf=pdfs[0], registration_file=reg_files[0], ignored=ignored)


# ---- Credentials ---------------------------------------------------------


@dataclass
class Credentials:
    username: str
    password: str
    env: str = "QA"
    crm_base_url: str | None = None


_PLACEHOLDER_PREFIXES = ("<ADD", "<add")


def _looks_like_placeholder(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    s = value.strip()
    return s.startswith(_PLACEHOLDER_PREFIXES) or s.endswith(">")


def load_credentials(path: Path = CREDENTIALS_PATH, env: str | None = None) -> Credentials:
    """Load CRM credentials for the selected environment.

    Schema (current): top-level "Env" selects a sub-object whose name matches (case-insensitive).
        {
          "Env": "QA",
          "QA":         {"username": ..., "password": ..., "crm_base_url": ...},
          "Production": {"username": ..., "password": ..., "crm_base_url": ...}
        }
    Schema (legacy, still accepted): flat object with just username/password at the top.
        { "username": ..., "password": ... }

    `env` argument (case-insensitive) overrides the file's "Env" field.
    """
    if not path.exists():
        raise FileNotFoundError(
            f"credentials.json not found at {path}. Create it with QA/Production blocks (see project README)."
        )
    data = load_json(path)

    # Legacy flat shape — no env routing
    if "username" in data and "password" in data and not any(
        isinstance(v, dict) for v in data.values()
    ):
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""
        if not username or not password:
            raise ValueError("credentials.json is missing username or password.")
        return Credentials(username=username, password=password, env="QA")

    # New env-aware shape
    selected_env = (env or data.get("Env") or "QA").strip()
    block = None
    block_name_used = None
    for key, value in data.items():
        if not isinstance(value, dict):
            continue
        if key.lower() == selected_env.lower():
            block = value
            block_name_used = key
            break
    if block is None:
        available = [k for k, v in data.items() if isinstance(v, dict)]
        raise ValueError(
            f"credentials.json has no environment block named {selected_env!r}. "
            f"Available: {available}"
        )

    username = (block.get("username") or "").strip()
    password = block.get("password") or ""
    crm_base_url = (block.get("crm_base_url") or "").strip() or None

    missing = [n for n, v in (("username", username), ("password", password)) if not v]
    placeholders = [n for n, v in (("username", username), ("password", password), ("crm_base_url", crm_base_url))
                    if v is not None and _looks_like_placeholder(v)]
    if missing:
        raise ValueError(
            f"credentials.json env {block_name_used!r} is missing: {missing}"
        )
    if placeholders:
        raise ValueError(
            f"credentials.json env {block_name_used!r} still has placeholder values for: {placeholders}. "
            "Replace the <ADD ...> placeholders with real values before running this environment."
        )

    return Credentials(
        username=username,
        password=password,
        env=block_name_used,
        crm_base_url=crm_base_url,
    )


# ---- Run folder & timestamp ----------------------------------------------


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M%S")


def build_run_folder_name(po_number: str | None, *, failure: bool = False) -> str:
    ts = utc_timestamp()
    po = po_number or "UNKNOWN"
    return f"{ts}_PO{po}"


def make_run_folder(po_number: str | None, *, failure: bool = False) -> Path:
    base = FAILURES_DIR if failure else PROCESSED_DIR
    folder = base / build_run_folder_name(po_number, failure=failure)
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "screenshots").mkdir(exist_ok=True)
    return folder


# ---- String helpers -------------------------------------------------------


_WS_RE = re.compile(r"\s+")


def collapse_ws(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return _WS_RE.sub(" ", value).strip()


def title_case_manufacturer(value: str) -> str:
    """Apply the SOP rule: first letter uppercase, remaining lowercase, per whitespace-separated word."""
    if not isinstance(value, str):
        return value
    parts = value.strip().split()
    return " ".join(p[:1].upper() + p[1:].lower() for p in parts if p)


def normalize_yes_no(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    v = value.strip().lower()
    if v in ("yes", "y", "true", "1"):
        return "Yes"
    if v in ("no", "n", "false", "0"):
        return "No"
    return value


def normalize_name(value: str) -> str:
    """Lowercase, collapse whitespace — used for cross-file name comparison."""
    return _WS_RE.sub(" ", (value or "").strip().lower())


def looks_like_sku(value: Any, pattern: str | None = None) -> bool:
    """Heuristic: short alphanumeric mixed string with no spaces."""
    if not isinstance(value, str):
        return False
    v = value.strip()
    if " " in v:
        return False
    if pattern:
        return bool(re.match(pattern, v))
    # Default: at least one letter, at least one digit, length >= 6, hyphens allowed.
    return bool(re.match(r"^(?=.*[A-Za-z])(?=.*\d)[A-Za-z0-9-]{6,}$", v))


# ---- Date parsing helpers ------------------------------------------------


_DATE_FORMATS = ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%d/%m/%Y")


def parse_date_flexible(raw: Any) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def format_date_us(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.strftime("%m/%d/%Y")


def format_date_iso(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d")


# ---- Generic ---------------------------------------------------------------


def first(items: Iterable[Any], default: Any = None) -> Any:
    for item in items:
        return item
    return default


# ---- Reference file lookup -----------------------------------------------


def lookup_sku_by_price(price: float | None, tolerance: float = 0.005) -> tuple[str, str] | None:
    """Return (sku, product_name) from the reference XLSX whose Sales price matches `price`.

    Scans the active sheet for a header row containing "SKU" and "Sales price" columns, then
    walks the data rows beneath it. Matches within `tolerance` to absorb floating-point rounding.
    Returns None if the file is missing, the price is None, or no match is found.
    """
    if price is None:
        return None

    ref_path = REFERENCES_DIR / _SKU_LIST_FILENAME
    if not ref_path.exists():
        return None

    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(ref_path), read_only=True, data_only=True)
        ws = wb.active

        sku_col = price_col = name_col = None
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if not isinstance(cell, str):
                    continue
                label = cell.strip().lower()
                if label == "sku":
                    sku_col = i
                elif label == "sales price":
                    price_col = i
                elif label.startswith("product/service") or label == "product name":
                    name_col = i
            if sku_col is not None and price_col is not None:
                break

        if sku_col is None or price_col is None:
            return None
        if name_col is None:
            name_col = 0  # product name is conventionally in column A

        # Re-iterate from the top, skipping rows until past the header. We identify "past header"
        # by waiting for a row whose price cell is numeric.
        for row in ws.iter_rows(values_only=True):
            if len(row) <= max(sku_col, price_col, name_col):
                continue
            sku = row[sku_col]
            sales_price = row[price_col]
            product_name = row[name_col]
            if not isinstance(sales_price, (int, float)):
                continue
            if sku is None:
                continue
            if abs(float(sales_price) - float(price)) <= tolerance:
                return (str(sku).strip(), str(product_name).strip() if product_name else "")
    except Exception:
        pass

    return None
