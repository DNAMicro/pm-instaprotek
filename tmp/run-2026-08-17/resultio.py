import openpyxl, re
F="/home/farsheed/pm-instaprotek/Insta-testing/Regression_QA_Log_NULLNET-2026-08-17.xlsx"

def load(tab):
    """Return (wb, ws, index) where index maps 'Section|ScenarioID' -> row."""
    wb=openpyxl.load_workbook(F); ws=wb[tab]; idx={}; sec=None
    for r in range(2, ws.max_row+1):
        a=ws.cell(r,1).value; b=ws.cell(r,2).value
        if a and str(a).strip(): sec=str(a).strip()
        if b is None: continue
        try: sid=int(float(str(b).strip()))
        except: continue
        idx[f"{sec}|{sid}"]=r
    return wb, ws, idx

def write(tab, results, defects=None):
    """results: {'Section|ID': (STATUS, 'actual result narrative')}"""
    wb, ws, idx = load(tab); wrote=[]; missed=[]
    for k,(st,note) in results.items():
        r=idx.get(k)
        if not r: missed.append(k); continue
        ws.cell(r,5).value=note
        ws.cell(r,6).value=st
        ws.cell(r,7).value=(defects or {}).get(k,"None")
        wrote.append(k)
    wb.save(F)
    return len(wrote), missed, idx

def tally(tab):
    wb, ws, idx = load(tab); c={}
    for k,r in idx.items():
        s=str(ws.cell(r,6).value or "NOT RUN").upper().strip()
        c[s]=c.get(s,0)+1
    return c, len(idx)

def sections(tab):
    wb, ws, idx = load(tab); out={}
    for k in idx:
        s=k.split("|")[0]; out[s]=out.get(s,0)+1
    return out
